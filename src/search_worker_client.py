import dataclasses
import time
import traceback
from datetime import datetime

import zerorpc
from openpyxl import load_workbook

from src.model import SearchParams, FoundCell
from src.ui.search_xls_rpc import SearchXlsServerRpcService
from src.utils import cell_value_match


@dataclasses.dataclass
class SearchClientWorker(SearchXlsServerRpcService):
    client: SearchXlsServerRpcService

    def __init__(self, port, process_id):
        self.client = zerorpc.Client()
        print("start rpc client " + str(port))
        self.server_url = "tcp://127.0.0.1:" + str(port)
        self.client.connect(self.server_url)
        # self.client.test_rpc("test_rpc")
        self.process_id = process_id

    def to_server_error(self, text):
        print("rpc client error : " + text)
        self.client.to_server_error(text)

    def to_server_report_search_finished(self, text):
        print("rpc client error : " + text)
        self.client.to_server_report_search_finished(text)

    def to_server_log(self, text):
        print("rpc client log : " + text)
        self.client.to_server_log(text)

    def to_server_report_finsh_one_file(self, file):
        self.client.to_server_report_finsh_one_file(file)

    def to_server_report_search_result(self, found_cell):
        self.client.to_server_report_search_result(found_cell.__dict__)

    def stop(self):
        self.client.to_server_report_search_finished(self.process_id)
        print("stop client ... ")
        self.client.disconnect(self.server_url)
        time.sleep(0.2)
        self.client.close()
        print("stop client ... done")

    def to_server_poll_file(self):
        return self.client.to_server_poll_file()

    def do_search_openpyxl(self, params: SearchParams, files):
        print("do_search_openpyxl, files = " + str(files))
        time.sleep(0.1)
        search_file_name = ""
        search_sheet_name = ""
        cell_name = ""
        try:
            while True:
                file = self.to_server_poll_file()
                if file is None:
                    print("no more files")
                    break
                start_time = datetime.now()
                workbook = load_workbook(filename=file, read_only=True)
                search_file_name = file
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    search_sheet_name = sheet
                    # sheet名字中包含, 则写入到FoundCell
                    if cell_value_match(
                            sheet,
                            params.search_text,
                            params.is_strict,
                            params.match_case,
                    ):
                        self.to_server_report_search_result(
                            FoundCell(
                                path=file,
                                sheet=sheet,
                                cell="",
                                cell_value="",
                            )
                        )
                    if params.search_text and not params.search_text.isspace():
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell_value = cell.value
                                if cell_value is not None and cell_value_match(
                                        cell_value,
                                        params.search_text,
                                        params.is_strict,
                                        params.match_case,
                                ):
                                    row_all_values = []
                                    for row_item in row:
                                        if row_item.value is not None:
                                            row_all_values.append(row_item.value)
                                            # row_all_values = row_all_values + str(row_item.value) + " | "
                                    self.to_server_report_search_result(
                                        FoundCell(
                                            path=file,
                                            sheet=sheet,
                                            cell=cell.coordinate,
                                            cell_value=cell_value,
                                            row_all=row_all_values,
                                        )
                                    )
                end_time = datetime.now()
                self.to_server_report_finsh_one_file(f"{file} : 耗时 : {str(end_time - start_time)}")
        except Exception as e:
            traceback.print_tb(e.__traceback__)
            self.to_server_error(str(e))
            print("err : ", search_file_name, search_sheet_name, str(e))
        finally:
            pass
